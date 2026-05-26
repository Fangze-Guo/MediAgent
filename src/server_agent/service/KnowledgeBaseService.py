"""
知识库服务层
处理知识库和文档相关的业务逻辑
"""
import asyncio
import logging
import mimetypes
import pathlib
import shutil
from typing import List

from fastapi import UploadFile

from src.server_agent.constants.CommonConstants import DATASET_PATH
from src.server_agent.exceptions import (
    NotFoundError,
    ServiceError,
    ValidationError,
    handle_service_exception,
)
from src.server_agent.mapper.KnowledgeBaseMapper import KnowledgeBaseMapper
from src.server_agent.model.entity.KbDocumentInfo import KbDocumentInfo
from src.server_agent.model.entity.KnowledgeBaseInfo import KnowledgeBaseInfo
from src.server_agent.model.vo.KnowledgeBaseVO import KbDocumentVO, KnowledgeBaseVO
from src.server_agent.service.EmbeddingService import EmbeddingService

logger = logging.getLogger(__name__)

ALLOWED_EXTENSIONS = {
    ".pdf", ".docx", ".doc", ".xlsx", ".xls", ".txt", ".md", ".csv"
}

CONTENT_TYPE_MAP = {
    ".pdf": "application/pdf",
    ".docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    ".doc": "application/msword",
    ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    ".xls": "application/vnd.ms-excel",
    ".txt": "text/plain",
    ".md": "text/markdown",
    ".csv": "text/csv",
}


class KnowledgeBaseService:
    """知识库服务类"""

    def __init__(self):
        self.mapper = KnowledgeBaseMapper()
        self.embedding = EmbeddingService()

    # ==================== 知识库 CRUD ====================

    @handle_service_exception
    async def create_kb(self, name: str, description: str = None, created_by: int = None) -> KnowledgeBaseVO:
        """创建知识库"""
        if not name or not name.strip():
            raise ValidationError(detail="知识库名称不能为空", context={"name": name})

        existing = await self.mapper.get_kb_by_name(name.strip())
        if existing:
            raise ValidationError(detail=f"知识库 '{name}' 已存在", context={"name": name})

        kb = KnowledgeBaseInfo(name=name.strip(), description=description, created_by=created_by)
        kb_id = await self.mapper.create_kb(kb)
        kb.id = kb_id

        return KnowledgeBaseVO(
            id=kb_id,
            name=kb.name,
            description=kb.description,
            documents=[],
            document_count=0,
        )

    @handle_service_exception
    async def get_kb_list(self) -> List[KnowledgeBaseVO]:
        """获取全部知识库列表（所有用户可见）"""
        kbs = await self.mapper.get_all_kbs()
        result = []
        for kb in kbs:
            docs = await self.mapper.get_documents_by_kb(kb.id)
            doc_vos = [self._doc_to_vo(d) for d in docs]
            result.append(KnowledgeBaseVO(
                id=kb.id,
                name=kb.name,
                description=kb.description,
                documents=doc_vos,
                document_count=len(doc_vos),
                chunk_count=sum(d.chunk_count for d in docs),
                created_at=kb.created_at,
                updated_at=kb.updated_at,
            ))
        return result

    @handle_service_exception
    async def get_kb_by_id(self, kb_id: int) -> KnowledgeBaseVO:
        """获取知识库详情（含文档列表）"""
        kb = await self.mapper.get_kb_by_id(kb_id)
        if not kb:
            raise NotFoundError(detail="知识库不存在", context={"kb_id": kb_id})

        docs = await self.mapper.get_documents_by_kb(kb_id)
        doc_vos = [self._doc_to_vo(d) for d in docs]
        return KnowledgeBaseVO(
            id=kb.id,
            name=kb.name,
            description=kb.description,
            documents=doc_vos,
            document_count=len(doc_vos),
            chunk_count=sum(d.chunk_count for d in docs),
            created_at=kb.created_at,
            updated_at=kb.updated_at,
        )

    @handle_service_exception
    async def update_kb(self, kb_id: int, name: str = None, description: str = None) -> KnowledgeBaseVO:
        """更新知识库信息（仅 admin）"""
        kb = await self.mapper.get_kb_by_id(kb_id)
        if not kb:
            raise NotFoundError(detail="知识库不存在", context={"kb_id": kb_id})

        if name and name.strip() != kb.name:
            existing = await self.mapper.get_kb_by_name(name.strip())
            if existing and existing.id != kb_id:
                raise ValidationError(detail=f"知识库名称 '{name}' 已被使用", context={"name": name})

        update_data = {}
        if name is not None:
            update_data["name"] = name.strip()
        if description is not None:
            update_data["description"] = description

        await self.mapper.update_kb(kb_id, update_data)
        return await self.get_kb_by_id(kb_id)

    @handle_service_exception
    async def delete_kb(self, kb_id: int) -> bool:
        """删除知识库及其所有文件（仅 admin）"""
        kb = await self.mapper.get_kb_by_id(kb_id)
        if not kb:
            raise NotFoundError(detail="知识库不存在", context={"kb_id": kb_id})

        # 删除磁盘文件夹
        kb_dir = pathlib.Path(DATASET_PATH) / "public" / "knowledge_base" / str(kb_id)
        try:
            if kb_dir.exists():
                shutil.rmtree(kb_dir)
                logger.info(f"Deleted KB folder: {kb_dir}")
        except Exception as e:
            logger.error(f"Failed to delete KB folder: {e}")

        await self.mapper.delete_kb(kb_id)
        await self.embedding.delete_kb_embeddings(kb_id)
        return True

    # ==================== 文档操作 ====================

    @handle_service_exception
    async def upload_documents(self, kb_id: int, files: List[UploadFile]) -> List[KbDocumentVO]:
        """上传文档到知识库（仅 admin）"""
        kb = await self.mapper.get_kb_by_id(kb_id)
        if not kb:
            raise NotFoundError(detail="知识库不存在", context={"kb_id": kb_id})

        kb_dir = pathlib.Path(DATASET_PATH) / "public" / "knowledge_base" / str(kb_id)
        kb_dir.mkdir(parents=True, exist_ok=True)

        uploaded = []
        for file in files:
            suffix = pathlib.Path(file.filename).suffix.lower()
            if suffix not in ALLOWED_EXTENSIONS:
                logger.warning(f"Skipped unsupported file: {file.filename}")
                continue

            content_type = CONTENT_TYPE_MAP.get(suffix) or (mimetypes.guess_type(file.filename)[0] or "application/octet-stream")

            # 处理同名文件
            target = kb_dir / file.filename
            if target.exists():
                stem = target.stem
                ext = target.suffix
                counter = 1
                while target.exists():
                    target = kb_dir / f"{stem}_{counter}{ext}"
                    counter += 1

            content = await file.read()
            with open(target, "wb") as f:
                f.write(content)

            rel_path = f"public/knowledge_base/{kb_id}/{target.name}"
            doc = KbDocumentInfo(
                kb_id=kb_id,
                file_name=target.name,
                file_path=rel_path,
                file_size=len(content),
                content_type=content_type,
                status="uploaded",
            )
            doc_id = await self.mapper.create_document(doc)
            doc.id = doc_id
            uploaded.append(self._doc_to_vo(doc))
            logger.info(f"Uploaded document: {target}")

        return uploaded

    async def _embed_and_update(
        self,
        doc_id: int,
        kb_id: int,
        file_path: pathlib.Path,
        chunk_size: int = None,
        chunk_overlap: int = None,
    ) -> None:
        """后台任务：embedding 完成后更新状态和 chunk 数。"""
        try:
            kwargs = {}
            if chunk_size is not None:
                kwargs["chunk_size"] = chunk_size
            if chunk_overlap is not None:
                kwargs["chunk_overlap"] = chunk_overlap
            count = await self.embedding.embed_document(doc_id, kb_id, file_path, **kwargs)
            await self.mapper.update_document_chunk_count(doc_id, count)
            await self.mapper.update_document_status(doc_id, "completed")
            logger.info(f"Embedding done: doc {doc_id}, {count} chunks")
        except Exception as exc:
            logger.error(f"Embedding failed for doc {doc_id}: {exc}")
            await self.mapper.update_document_status(doc_id, "failed")

    @handle_service_exception
    async def analyze_document(
        self,
        kb_id: int,
        doc_id: int,
        chunk_size: int = None,
        chunk_overlap: int = None,
    ) -> KbDocumentVO:
        """手动触发文档 embedding（先清旧向量，再后台生成）。"""
        doc = await self.mapper.get_document_by_id(doc_id)
        if not doc:
            raise NotFoundError(detail="文档不存在", context={"doc_id": doc_id})
        if doc.kb_id != kb_id:
            raise ValidationError(detail="文档不属于该知识库", context={"doc_id": doc_id, "kb_id": kb_id})

        file_path = pathlib.Path(DATASET_PATH) / doc.file_path
        if not file_path.exists():
            raise ServiceError(detail="文件不存在于磁盘", context={"path": str(file_path)})

        await self.embedding.delete_document_embeddings(doc_id, kb_id)
        await self.mapper.update_document_status(doc_id, "processing")
        asyncio.create_task(
            self._embed_and_update(doc_id, kb_id, file_path, chunk_size, chunk_overlap)
        )
        doc.status = "processing"
        return self._doc_to_vo(doc)

    @handle_service_exception
    async def delete_document(self, kb_id: int, doc_id: int) -> bool:
        """删除文档（仅 admin）"""
        doc = await self.mapper.get_document_by_id(doc_id)
        if not doc:
            raise NotFoundError(detail="文档不存在", context={"doc_id": doc_id})
        if doc.kb_id != kb_id:
            raise ValidationError(detail="文档不属于该知识库", context={"doc_id": doc_id, "kb_id": kb_id})

        # 删除磁盘文件
        file_path = pathlib.Path(DATASET_PATH) / doc.file_path
        try:
            if file_path.exists():
                file_path.unlink()
                logger.info(f"Deleted document file: {file_path}")
        except Exception as e:
            logger.error(f"Failed to delete document file: {e}")

        await self.embedding.delete_document_embeddings(doc_id, kb_id)
        await self.mapper.delete_document(doc_id)
        return True

    @handle_service_exception
    async def search_kb(self, kb_id: int, query: str, top_k: int = 5) -> List[dict]:
        """语义检索知识库，返回最相关的 top_k 个 chunk。"""
        kb = await self.mapper.get_kb_by_id(kb_id)
        if not kb:
            raise NotFoundError(detail="知识库不存在", context={"kb_id": kb_id})
        return await self.embedding.search(kb_id, query, top_k)

    @handle_service_exception
    async def get_document_preview(self, kb_id: int, doc_id: int) -> dict:
        """
        获取文档预览信息：
        - PDF/图片：返回 serve_url 供前端直接嵌入
        - Word：提取纯文本返回
        - Excel：返回 JSON 表格数据
        - TXT/MD/CSV：返回原始文本
        """
        doc = await self.mapper.get_document_by_id(doc_id)
        if not doc:
            raise NotFoundError(detail="文档不存在", context={"doc_id": doc_id})
        if doc.kb_id != kb_id:
            raise ValidationError(detail="文档不属于该知识库", context={"doc_id": doc_id, "kb_id": kb_id})

        suffix = pathlib.Path(doc.file_name).suffix.lower()
        file_path = pathlib.Path(DATASET_PATH) / doc.file_path

        if not file_path.exists():
            raise ServiceError(detail="文件不存在于磁盘", context={"path": str(file_path)})

        # PDF — 直接给前端 URL，浏览器内嵌预览
        if suffix == ".pdf":
            from urllib.parse import quote
            encoded_path = quote(doc.file_path, safe="/")
            return {
                "type": "url",
                "serve_url": f"/files/serve/{encoded_path}",
                "file_name": doc.file_name,
                "content_type": doc.content_type,
            }

        # Word
        if suffix in (".docx", ".doc"):
            text = self._extract_word_text(file_path)
            return {"type": "text", "content": text, "file_name": doc.file_name}

        # Excel
        if suffix in (".xlsx", ".xls"):
            data = self._extract_excel_data(file_path)
            return {"type": "table", "sheets": data, "file_name": doc.file_name}

        # TXT / MD / CSV — 直接读取
        try:
            text = file_path.read_text(encoding="utf-8", errors="replace")
            return {"type": "text", "content": text, "file_name": doc.file_name}
        except Exception as e:
            raise ServiceError(detail="读取文件失败", context={"error": str(e)})

    # ==================== 内部工具 ====================

    def _doc_to_vo(self, doc: KbDocumentInfo) -> KbDocumentVO:
        return KbDocumentVO(
            id=doc.id,
            file_name=doc.file_name,
            file_path=doc.file_path,
            file_size=doc.file_size,
            content_type=doc.content_type,
            knowledge_base_id=doc.kb_id,
            status=doc.status,
            created_at=doc.created_at,
            updated_at=doc.updated_at,
            processing_tasks=[{
                "id": doc.id,
                "status": doc.status,
                "error_message": None,
                "created_at": doc.created_at,
                "updated_at": doc.updated_at,
            }],
        )

    def _extract_word_text(self, file_path: pathlib.Path) -> str:
        try:
            from docx import Document
            doc = Document(str(file_path))
            return "\n".join(p.text for p in doc.paragraphs if p.text.strip())
        except ImportError:
            raise ServiceError(detail="Word 预览需要 python-docx，请先安装", context={})
        except Exception as e:
            raise ServiceError(detail="解析 Word 文档失败", context={"error": str(e)})

    def _extract_excel_data(self, file_path: pathlib.Path) -> dict:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(str(file_path), read_only=True, data_only=True)
            sheets = {}
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = []
                for row in ws.iter_rows(values_only=True):
                    if any(cell is not None for cell in row):
                        rows.append([str(cell) if cell is not None else "" for cell in row])
                sheets[sheet_name] = rows
            wb.close()
            return sheets
        except ImportError:
            raise ServiceError(detail="Excel 预览需要 openpyxl，请先安装", context={})
        except Exception as e:
            raise ServiceError(detail="解析 Excel 文档失败", context={"error": str(e)})
