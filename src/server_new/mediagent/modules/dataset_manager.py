from __future__ import annotations
"""
dataset_manager.py —— 数据集管理 & 摘要模块（异步）

适配你现在的 SQLite 表结构：
CREATE TABLE dataset_catalog (
    dataset_name            TEXT,
    case_count              INTEGER,
    clinical_data_desc      TEXT,
    text_data_desc          TEXT,
    imaging_data_desc       TEXT,
    pathology_data_desc     TEXT,
    genomics_data_desc      TEXT,
    annotation_desc         TEXT,
    notes                   TEXT,
    user_id                 INTEGER,
    id                      INTEGER PRIMARY KEY,   -- 唯一数据集ID
    has_data                INT DEFAULT 0,
    has_description_file    INT DEFAULT 0,
    data_path               TEXT,                  -- 相对 DATASET_STORAGE_ROOT 的目录
    create_time             TEXT DEFAULT datetime('now','localtime'),
    description_file_path   TEXT                   -- 相对 DATASET_STORAGE_ROOT 的具体描述文件路径
);

关键注意点：
1. 权限模型：
   - user_id == -1   → 公共数据集，所有用户可见
   - user_id == X    → 仅该用户可见
2. dataset_id == id 作为后续 API 的唯一引用
3. data_path / description_file_path 是相对路径，实际磁盘地址 = DATASET_STORAGE_ROOT / <相对路径>

对外主要接口：
- overview(user_uid, db_path?, limit?) -> {"ok": bool, "text": str, "items": [...]}
- focus(dataset_id, user_need_text, user_uid, db_path?) -> {"ok": bool, "text": str}

focus()：
- 会校验权限
- 会读取 description_file_path 指向的 CSV / XLSX
- 会尝试调用 LLM 生成简介；过大则 map-reduce；LLM 不可用就走本地估计
"""

import csv
import io
import json
import os
import sqlite3
import asyncio
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from src.server_new.mediagent.paths import in_data

# ===================== 统一路径基准 =====================
# 这个根目录是所有 data_path / description_file_path 的相对基准
DATASET_STORAGE_ROOT: Path = in_data("files")

# SQLite 数据库路径
DATASET_DB_PATH: Path = in_data("db", "app.sqlite3")

# （兼容旧逻辑）公共信息文件可能放的目录，用于兜底 locate_file()
DATASET_FILES_ROOT: Path = in_data("files", "public", "dataset_info")

# ===================== LLM 网关配置（可按需改） =====================
LLM_MODEL: Optional[str] = "deepseek-chat"
LLM_BASE_URL: Optional[str] = "https://api.deepseek.com/v1"
LLM_API_KEY: Optional[str] = "sk-d0e27c4c590a454e8284309067c03f04"
LLM_TIMEOUT: float = 60.0

# ===================== 读取与大小控制 =====================
MAX_TEXT_BYTES_FOR_LLM: int = 2 * 1024 * 1024
CSV_ENCODING_CANDIDATES: Tuple[str, ...] = ("utf-8-sig", "utf-8", "gbk", "latin-1")
SAMPLE_HEAD_ROWS: int = 200
LOCAL_SUMMARY_ROWS: int = 50

# Map-Reduce 切块策略
SINGLEPASS_MAX_BYTES: int = 512 * 1024      # 单轮直接丢给 LLM 的上限
CHUNK_TARGET_BYTES: int = 256 * 1024        # 每块目标大小
CHUNK_HARD_MAX_BYTES: int = 320 * 1024      # 每块硬上限

# ===================== 数据类 =====================

@dataclass
class CatalogRow:
    id: int
    user_id: int
    dataset_name: str
    case_count: Optional[int]

    clinical_data_desc: Optional[str]
    text_data_desc: Optional[str]
    imaging_data_desc: Optional[str]
    pathology_data_desc: Optional[str]
    genomics_data_desc: Optional[str]
    annotation_desc: Optional[str]
    notes: Optional[str]

    has_data: int
    has_description_file: int
    data_path: Optional[str]
    description_file_path: Optional[str]
    # create_time 存在于表里，但当前逻辑没有强依赖它。
    # 如果以后需要展示创建时间，可以加进来。


# ===================== 对外接口（异步） =====================

async def overview(
    *,
    user_uid: int,
    db_path: Optional[Path] = None,
    limit: Optional[int] = None,
) -> Dict[str, Any]:
    """
    概览调用者 user_uid 可见的数据集（含公共 user_id = -1）。
    返回:
    {
      "ok": True,
      "text": "...若干行中文摘要...",
      "items": [
        {
          "id": <dataset_id:int>,
          "name": <dataset_name:str>,
          "case_count": <int|None>,
          "has_data": 0/1,
          "has_description_file": 0/1,
          "is_public": bool
        },
        ...
      ]
    }
    """
    try:
        rows = await _fetch_catalog_rows_async(
            db_path or DATASET_DB_PATH,
            user_uid=user_uid,
            limit=limit,
        )
        if not rows:
            return {
                "ok": True,
                "text": "当前数据库中没有可用的数据集条目。",
                "items": []
            }

        parts: List[str] = [f"共收录可访问数据集 {len(rows)} 个。"]
        items: List[Dict[str, Any]] = []

        for i, r in enumerate(rows, 1):
            cc_str = f"{r.case_count}例" if isinstance(r.case_count, int) else "例数未知"
            im = _short(r.imaging_data_desc)
            tx = _short(r.text_data_desc)
            pa = _short(r.pathology_data_desc)
            ge = _short(r.genomics_data_desc)
            an = _short(r.annotation_desc)

            parts.append(
                f"{i}. {r.dataset_name}：{cc_str}；影像[{im}]；文本[{tx}]；"
                f"病理[{pa}]；基因[{ge}]；标注[{an}]。"
            )

            items.append({
                "id": r.id,
                "name": r.dataset_name,
                "case_count": r.case_count,
                "has_data": r.has_data,
                "has_description_file": r.has_description_file,
                "is_public": (r.user_id == -1),
            })

        return {
            "ok": True,
            "text": "\n".join(parts),
            "items": items
        }

    except Exception as e:
        return {"ok": False, "text": f"读取 catalog 失败：{e!r}", "items": []}


async def focus(
    dataset_id: int,
    user_need_text: str,
    *,
    user_uid: int,
    db_path: Optional[Path] = None,
) -> Dict[str, Any]:
    """
    针对单个数据集做定点查询并生成短摘要。

    步骤：
    1. 根据 dataset_id 精确查询该条记录。
    2. 权限检查：
       - user_id == -1 → 公共
       - user_id == user_uid → 私有但属于当前用户
       否则拒绝。
    3. 检查 has_description_file / description_file_path。
    4. 读取该描述文件 (csv/xlsx) 转成纯文本。
    5. 用 LLM(summary) 或 Map-Reduce。LLM不可用时走本地摘要。

    返回：
    { "ok": True, "text": "..." }
    or
    { "ok": False, "text": "出错原因" }
    """
    try:
        # 1) 取该条记录
        row = await _fetch_single_row_async(
            db_path or DATASET_DB_PATH,
            dataset_id=dataset_id,
        )
        if row is None:
            return {"ok": False, "text": "未找到该数据集。"}

        # 2) 权限检查
        if (row.user_id != -1) and (row.user_id != user_uid):
            return {"ok": False, "text": "无权访问该数据集。"}

        # 3) 检查描述文件可用性
        if row.has_description_file != 1 or not row.description_file_path:
            return {
                "ok": False,
                "text": "该数据集尚未提供描述文件，无法生成摘要。"
            }

        # 4) 解析绝对路径并读取
        desc_abs = _resolve_rel_path(row.description_file_path)
        if not desc_abs.exists():
            return {
                "ok": False,
                "text": f"描述文件不存在或已被移除：{row.description_file_path}"
            }

        fmt = _guess_format_from_suffix(desc_abs)
        if fmt == "csv":
            full_text, _trunc, file_info = await asyncio.to_thread(_read_csv_full_text, desc_abs)
        elif fmt == "xlsx":
            full_text, _trunc, file_info = await asyncio.to_thread(_read_xlsx_full_tsv, desc_abs)
        else:
            # 兜底尝试旧逻辑 locate_file(dataset_name)
            alt_path, alt_fmt = await asyncio.to_thread(_locate_file, row.dataset_name)
            if alt_path is None:
                return {"ok": False, "text": "描述文件类型未知或无法读取。"}
            if alt_fmt == "csv":
                full_text, _trunc, file_info = await asyncio.to_thread(_read_csv_full_text, alt_path)
            else:
                full_text, _trunc, file_info = await asyncio.to_thread(_read_xlsx_full_tsv, alt_path)

        dataset_meta = {
            "dataset_id": row.id,
            "dataset_name": row.dataset_name,
            "case_count": row.case_count,
            "has_data": row.has_data,
            "has_description_file": row.has_description_file,
        }

        total_bytes = len(full_text.encode("utf-8"))

        # 5a) LLM 可用，且文本不大 → 单轮摘要
        if await _llm_available_async():
            if total_bytes <= SINGLEPASS_MAX_BYTES:
                user_prompt = _build_user_prompt_for_llm(
                    dataset_meta=dataset_meta,
                    user_need_text=user_need_text,
                    file_text=full_text,
                    truncated=False,
                    file_info={**file_info, "pipeline": "single-pass", "bytes": total_bytes},
                )
                out = await _call_llm_async(_fixed_system_prompt(), user_prompt)
                if out:
                    return {"ok": True, "text": out.strip()}
            else:
                # 5b) LLM 可用，文本很大 → Map-Reduce
                merged = await _map_reduce_summarize(
                    system_prompt=_fixed_system_prompt(),
                    dataset_meta=dataset_meta,
                    user_need_text=user_need_text,
                    full_text=full_text,
                    base_file_info=file_info,
                )
                if merged:
                    return {"ok": True, "text": merged.strip()}

        # 5c) LLM 不可用或失败 → fallback 本地摘要
        local_text = await asyncio.to_thread(
            _local_summarize,
            dataset_meta,
            user_need_text,
            full_text,
            file_info,
            LOCAL_SUMMARY_ROWS,
        )
        return {"ok": True, "text": local_text}

    except Exception as e:
        return {"ok": False, "text": f"处理失败：{e!r}"}


# ===================== DB 查询（优先 aiosqlite） =====================

async def _fetch_catalog_rows_async(
    db_path: Path,
    *,
    user_uid: int,
    limit: Optional[int],
) -> List[CatalogRow]:
    """
    返回当前 user_uid 可见的数据集（包括公共 user_id=-1）。
    """
    db_path = Path(db_path).expanduser().resolve()
    if not db_path.exists():
        raise FileNotFoundError(f"catalog 数据库不存在：{db_path}")

    sql = (
        "SELECT "
        "id, "
        "user_id, "
        "dataset_name, "
        "case_count, "
        "clinical_data_desc, "
        "text_data_desc, "
        "imaging_data_desc, "
        "pathology_data_desc, "
        "genomics_data_desc, "
        "annotation_desc, "
        "notes, "
        "has_data, "
        "has_description_file, "
        "data_path, "
        "description_file_path "
        "FROM dataset_catalog "
        "WHERE user_id = ? OR user_id = -1 "
        "ORDER BY dataset_name ASC"
    )

    try:
        import aiosqlite  # type: ignore
    except Exception:
        return await asyncio.to_thread(
            _fetch_catalog_rows_sync,
            db_path,
            user_uid,
            limit
        )

    rows: List[CatalogRow] = []
    async with aiosqlite.connect(db_path.as_posix()) as conn:
        async with conn.execute(sql, (user_uid,)) as cur:
            i = 0
            async for (
                rid, uid, dn, cc,
                cl, tx, im, pa,
                ge, an, no,
                hd, hf, dp, dfp
            ) in cur:
                rows.append(
                    CatalogRow(
                        id=int(rid),
                        user_id=int(uid),
                        dataset_name=str(dn),
                        case_count=int(cc) if cc is not None else None,
                        clinical_data_desc=_nz(cl),
                        text_data_desc=_nz(tx),
                        imaging_data_desc=_nz(im),
                        pathology_data_desc=_nz(pa),
                        genomics_data_desc=_nz(ge),
                        annotation_desc=_nz(an),
                        notes=_nz(no),
                        has_data=int(hd) if hd is not None else 0,
                        has_description_file=int(hf) if hf is not None else 0,
                        data_path=_nz(dp),
                        description_file_path=_nz(dfp),
                    )
                )
                i += 1
                if limit is not None and i >= limit:
                    break
    return rows


def _fetch_catalog_rows_sync(
    db_path: Path,
    user_uid: int,
    limit: Optional[int],
) -> List[CatalogRow]:
    sql = (
        "SELECT "
        "id, "
        "user_id, "
        "dataset_name, "
        "case_count, "
        "clinical_data_desc, "
        "text_data_desc, "
        "imaging_data_desc, "
        "pathology_data_desc, "
        "genomics_data_desc, "
        "annotation_desc, "
        "notes, "
        "has_data, "
        "has_description_file, "
        "data_path, "
        "description_file_path "
        "FROM dataset_catalog "
        "WHERE user_id = ? OR user_id = -1 "
        "ORDER BY dataset_name ASC"
    )

    rows: List[CatalogRow] = []
    with sqlite3.connect(db_path.as_posix(), check_same_thread=False) as conn:
        cur = conn.execute(sql, (user_uid,))
        for i, (
            rid, uid, dn, cc,
            cl, tx, im, pa,
            ge, an, no,
            hd, hf, dp, dfp
        ) in enumerate(cur):
            rows.append(
                CatalogRow(
                    id=int(rid),
                    user_id=int(uid),
                    dataset_name=str(dn),
                    case_count=int(cc) if cc is not None else None,
                    clinical_data_desc=_nz(cl),
                    text_data_desc=_nz(tx),
                    imaging_data_desc=_nz(im),
                    pathology_data_desc=_nz(pa),
                    genomics_data_desc=_nz(ge),
                    annotation_desc=_nz(an),
                    notes=_nz(no),
                    has_data=int(hd) if hd is not None else 0,
                    has_description_file=int(hf) if hf is not None else 0,
                    data_path=_nz(dp),
                    description_file_path=_nz(dfp),
                )
            )
            if limit is not None and i + 1 >= limit:
                break
    return rows


async def _fetch_single_row_async(
    db_path: Path,
    *,
    dataset_id: int,
) -> Optional[CatalogRow]:
    """
    根据主键 id 取到单行 CatalogRow。
    """
    db_path = Path(db_path).expanduser().resolve()
    if not db_path.exists():
        raise FileNotFoundError(f"catalog 数据库不存在：{db_path}")

    sql = (
        "SELECT "
        "id, "
        "user_id, "
        "dataset_name, "
        "case_count, "
        "clinical_data_desc, "
        "text_data_desc, "
        "imaging_data_desc, "
        "pathology_data_desc, "
        "genomics_data_desc, "
        "annotation_desc, "
        "notes, "
        "has_data, "
        "has_description_file, "
        "data_path, "
        "description_file_path "
        "FROM dataset_catalog "
        "WHERE id = ?"
    )

    try:
        import aiosqlite  # type: ignore
    except Exception:
        return await asyncio.to_thread(
            _fetch_single_row_sync, db_path, dataset_id
        )

    async with aiosqlite.connect(db_path.as_posix()) as conn:
        async with conn.execute(sql, (dataset_id,)) as cur:
            row = await cur.fetchone()
            if row is None:
                return None
            (
                rid, uid, dn, cc,
                cl, tx, im, pa,
                ge, an, no,
                hd, hf, dp, dfp
            ) = row
            return CatalogRow(
                id=int(rid),
                user_id=int(uid),
                dataset_name=str(dn),
                case_count=int(cc) if cc is not None else None,
                clinical_data_desc=_nz(cl),
                text_data_desc=_nz(tx),
                imaging_data_desc=_nz(im),
                pathology_data_desc=_nz(pa),
                genomics_data_desc=_nz(ge),
                annotation_desc=_nz(an),
                notes=_nz(no),
                has_data=int(hd) if hd is not None else 0,
                has_description_file=int(hf) if hf is not None else 0,
                data_path=_nz(dp),
                description_file_path=_nz(dfp),
            )


def _fetch_single_row_sync(
    db_path: Path,
    dataset_id: int,
) -> Optional[CatalogRow]:
    sql = (
        "SELECT "
        "id, "
        "user_id, "
        "dataset_name, "
        "case_count, "
        "clinical_data_desc, "
        "text_data_desc, "
        "imaging_data_desc, "
        "pathology_data_desc, "
        "genomics_data_desc, "
        "annotation_desc, "
        "notes, "
        "has_data, "
        "has_description_file, "
        "data_path, "
        "description_file_path "
        "FROM dataset_catalog "
        "WHERE id = ?"
    )
    with sqlite3.connect(db_path.as_posix(), check_same_thread=False) as conn:
        cur = conn.execute(sql, (dataset_id,))
        row = cur.fetchone()
        if row is None:
            return None
        (
            rid, uid, dn, cc,
            cl, tx, im, pa,
            ge, an, no,
            hd, hf, dp, dfp
        ) = row
        return CatalogRow(
            id=int(rid),
            user_id=int(uid),
            dataset_name=str(dn),
            case_count=int(cc) if cc is not None else None,
            clinical_data_desc=_nz(cl),
            text_data_desc=_nz(tx),
            imaging_data_desc=_nz(im),
            pathology_data_desc=_nz(pa),
            genomics_data_desc=_nz(ge),
            annotation_desc=_nz(an),
            notes=_nz(no),
            has_data=int(hd) if hd is not None else 0,
            has_description_file=int(hf) if hf is not None else 0,
            data_path=_nz(dp),
            description_file_path=_nz(dfp),
        )


# ===================== 文件路径与读取 =====================

def _resolve_rel_path(rel: str) -> Path:
    """
    将数据库中记录的相对路径（相对 DATASET_STORAGE_ROOT）转为绝对路径。
    例如:
        DATASET_STORAGE_ROOT = D:/.../data/files
        rel = "private/5931999430/dataset/测试数据集1/6068703125.csv"
    结果:
        D:/.../data/files/private/5931999430/dataset/测试数据集1/6068703125.csv
    """
    return (DATASET_STORAGE_ROOT / rel).resolve()


def _guess_format_from_suffix(p: Path) -> Optional[str]:
    suf = p.suffix.lower()
    if suf == ".csv":
        return "csv"
    if suf in (".xlsx", ".xlsm", ".xls"):
        return "xlsx"
    return None


def _locate_file(dataset_name: str) -> Tuple[Optional[Path], Optional[str]]:
    """
    旧版兜底逻辑（公共数据集场景）：
    在 DATASET_FILES_ROOT 下尝试 {dataset_name}.csv / .xlsx
    """
    root = DATASET_FILES_ROOT.expanduser().resolve()
    csv_path = root / f"{dataset_name}.csv"
    xlsx_path = root / f"{dataset_name}.xlsx"
    if csv_path.exists():
        return csv_path, "csv"
    if xlsx_path.exists():
        return xlsx_path, "xlsx"
    return None, None


def _detect_csv_encoding(p: Path) -> str:
    for enc in CSV_ENCODING_CANDIDATES:
        try:
            with p.open("r", encoding=enc) as f:
                f.read(1024)
            return enc
        except Exception:
            continue
    return "utf-8"


def _read_csv_full_text(p: Path) -> Tuple[str, bool, Dict[str, Any]]:
    enc = _detect_csv_encoding(p)
    txt = p.read_text(encoding=enc)
    return txt, False, {
        "format": "csv",
        "encoding": enc,
        "size_bytes": p.stat().st_size,
        "note": "full_text",
        "path": str(p),
    }


def _read_xlsx_full_tsv(p: Path) -> Tuple[str, bool, Dict[str, Any]]:
    """
    读取 XLSX 第一个 sheet，转成 TSV 风格的文本。
    """
    try:
        import openpyxl  # type: ignore
    except Exception:
        msg = f"(提示) 未安装 openpyxl，无法读取 {p.name} 的内容。请安装 openpyxl。"
        return msg, False, {
            "format": "xlsx",
            "error": "openpyxl_missing",
            "size_bytes": p.stat().st_size if p.exists() else None,
            "path": str(p),
        }

    from openpyxl import load_workbook  # type: ignore
    wb = load_workbook(p.as_posix(), read_only=True, data_only=True)
    sheetnames = wb.sheetnames
    if not sheetnames:
        try:
            wb.close()
        except Exception:
            pass
        return "(空表)", False, {"format": "xlsx", "sheets": 0, "path": str(p)}

    ws = wb[sheetnames[0]]
    out = io.StringIO()
    for row in ws.iter_rows(values_only=True):
        cells = [
            ("" if v is None else str(v)).replace("\r", " ").replace("\n", " ")
            for v in row
        ]
        out.write("\t".join(cells) + "\n")
    tsv = out.getvalue()
    try:
        wb.close()
    except Exception:
        pass
    return tsv, False, {
        "format": "xlsx",
        "sheet": sheetnames[0],
        "size_bytes": p.stat().st_size if p.exists() else None,
        "note": "full_text",
        "path": str(p),
    }


# ===================== LLM 支持 & 摘要降级 =====================

def _fixed_system_prompt() -> str:
    # 与之前保持一致
    return (
        "你是医疗数据助手。请根据给定的表格内容与用户需求，输出一段简短总结（100~200字），"
        "表格是对应同名数据集的信息表，用户想了解的是数据集的情况，你可以从表格出发进行合理推测。"
        "要求：1) 不要逐行罗列原始数据；2) 提炼列名、关键信息与可用性；3) 若信息不足，说明缺口；"
        "4) 中文输出；5) 不包含表格源码。"
    )


async def _llm_available_async() -> bool:
    if not LLM_MODEL:
        return False
    try:
        from openai import AsyncOpenAI  # type: ignore
    except Exception:
        return False
    return True


async def _call_llm_async(system_prompt: str, user_prompt: str) -> Optional[str]:
    try:
        from openai import AsyncOpenAI  # type: ignore
        client = AsyncOpenAI(
            api_key=(LLM_API_KEY or "lm-studio"),
            base_url=(LLM_BASE_URL or None),
            timeout=LLM_TIMEOUT,
        )
        resp = await client.chat.completions.create(
            model=LLM_MODEL,
            temperature=0.2,
            messages=[
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": user_prompt},
            ],
        )
        txt = (resp.choices[0].message.content or "").strip()
        return txt or None
    except Exception:
        return None


def _build_user_prompt_for_llm(
    *,
    dataset_meta: Dict[str, Any],
    user_need_text: str,
    file_text: str,
    truncated: bool,
    file_info: Dict[str, Any],
) -> str:
    meta = json.dumps(dataset_meta, ensure_ascii=False)
    info = json.dumps(file_info, ensure_ascii=False)
    tip = "（以下内容可能为采样片段）" if truncated else "（以下为完整内容或较完整片段）"
    return (
        f"【用户需求】\n{user_need_text}\n\n"
        f"【数据集元信息】\n{meta}\n\n"
        f"【文件信息】\n{info}\n\n"
        f"【表格文本】{tip}\n{file_text}"
    )


def _local_summarize(
    dataset_meta: Dict[str, Any],
    user_need_text: str,
    file_text: str,
    file_info: Dict[str, Any],
    max_rows: int = 50,
) -> str:
    """
    没有 LLM 或 LLM 失败时的兜底摘要。
    采样前 max_rows 行，推断列结构并给出高层总结。
    """
    lines = [ln for ln in (file_text or "").splitlines() if ln.strip()]
    head = lines[:max_rows]
    delimiter = "\t" if ("\t" in head[0] if head else False) else ","
    cols = head[0].split(delimiter) if head else []
    num_cols = len(cols)

    sample_hint = f"已采样前{len(head)}行" if head else "未能采样"
    info_hint = (
        f"格式 {file_info.get('format')}, "
        f"估计大小 {file_info.get('size_bytes','?')} 字节"
    )
    cols_short = ", ".join(cols[:10]) + ("..." if len(cols) > 10 else "")

    return (
        "数据集简要摘要（本地估计）:\n"
        f"- 数据集ID/名称：{dataset_meta.get('dataset_id')} / {dataset_meta.get('dataset_name')}\n"
        f"- 例数：{dataset_meta.get('case_count', '未知')}；"
        f"是否已有原始数据文件：{'是' if dataset_meta.get('has_data') else '否'}\n"
        f"- {info_hint}；{sample_hint}\n"
        f"- 估计列数：{num_cols}；部分列名：{cols_short if cols else '未知'}\n"
        f"- 用户需求：{user_need_text}\n"
        "若需更详细摘要，请配置 LLM 或提供更小范围的表格。"
    )


def _split_text_by_bytes_linesafe(text: str, target_bytes: int, hard_max: int) -> List[str]:
    """
    按字节大小切分字符串，尽量保证在行边界处断开。
    """
    lines = text.splitlines(keepends=True)
    chunks: List[str] = []
    buf: List[str] = []
    buf_bytes = 0
    for ln in lines:
        ln_b = len(ln.encode("utf-8"))
        # 如果加上这一行超了 target_bytes，就先把现有缓存落一块
        if buf and (buf_bytes + ln_b > target_bytes):
            # 特殊情况：单行本身就超过 hard_max 且当前块为空，只能硬切
            if buf_bytes == 0 and ln_b > hard_max:
                chunks.append(
                    ln[:hard_max].encode("utf-8", "ignore").decode("utf-8", "ignore")
                )
                rest = ln[hard_max:]
                if rest:
                    lines.insert(0, rest)  # 继续处理剩余片段
                continue
            # 正常结块
            chunks.append("".join(buf))
            buf, buf_bytes = [ln], ln_b
        else:
            buf.append(ln)
            buf_bytes += ln_b

        # 如果当前块已经接近 hard_max，不再继续塞
        if buf_bytes >= hard_max:
            chunks.append("".join(buf))
            buf, buf_bytes = [], 0

    if buf:
        chunks.append("".join(buf))
    return chunks


async def _map_reduce_summarize(
    *,
    system_prompt: str,
    dataset_meta: Dict[str, Any],
    user_need_text: str,
    full_text: str,
    base_file_info: Dict[str, Any],
) -> Optional[str]:
    """
    Map-Reduce 两阶段摘要：
    1. Map：按块调用同样的提示词，得到每块的摘要
    2. Reduce：把所有块摘要拼起来，再走同样提示词，得到最终合并摘要
    """
    chunks = _split_text_by_bytes_linesafe(full_text, CHUNK_TARGET_BYTES, CHUNK_HARD_MAX_BYTES)
    if not chunks:
        return None

    part_summaries: List[str] = []
    for idx, chunk in enumerate(chunks, start=1):
        file_info = {
            **base_file_info,
            "pipeline": "map-reduce",
            "stage": "map",
            "chunk_index": idx,
            "chunks_total": len(chunks),
            "bytes": len(chunk.encode("utf-8")),
        }
        user_prompt = _build_user_prompt_for_llm(
            dataset_meta=dataset_meta,
            user_need_text=user_need_text,
            file_text=chunk,
            truncated=False,
            file_info=file_info,
        )
        out = await _call_llm_async(system_prompt, user_prompt)
        if not out:
            continue
        part_summaries.append(f"[第{idx}/{len(chunks)}块摘要]\n{out.strip()}")

    if not part_summaries:
        return None

    merged_text = "\n\n".join(part_summaries)
    reduce_info = {
        **base_file_info,
        "pipeline": "map-reduce",
        "stage": "reduce",
        "chunks_total": len(chunks),
        "bytes": len(merged_text.encode("utf-8")),
    }
    final_user_prompt = _build_user_prompt_for_llm(
        dataset_meta=dataset_meta,
        user_need_text=user_need_text,
        file_text=merged_text,
        truncated=False,
        file_info=reduce_info,
    )
    final = await _call_llm_async(system_prompt, final_user_prompt)
    return final


# ===================== 小工具函数 =====================

def _nz(x: Optional[str]) -> Optional[str]:
    if x is None:
        return None
    s = str(x).strip()
    return s if s else None


def _short(x: Optional[str], n: int = 120) -> str:
    if not x:
        return "无"
    s = str(x).strip().replace("\n", " ")
    return (s[:n] + "…") if len(s) > n else s


# ===================== 命令行（异步调试入口） =====================

async def _amain(argv: Optional[List[str]] = None) -> int:
    """
    CLI 调试：
    1. overview:
       python dataset_manager.py overview --user 5931999430
    2. focus:
       python dataset_manager.py focus --id 6068703125 --user 5931999430 --need "总结一下数据内容"
    """
    import argparse
    p = argparse.ArgumentParser(description="dataset_manager（overview / focus, async）")
    sub = p.add_subparsers(dest="cmd", required=True)

    # overview
    p1 = sub.add_parser("overview", help="读取 catalog 并生成总览文本（按用户可见性过滤）")
    p1.add_argument("--db", type=str, default=str(DATASET_DB_PATH), help="SQLite DB 路径")
    p1.add_argument("--user", type=int, required=True, help="调用者 user_uid")
    p1.add_argument("--limit", type=int, default=None, help="最多读取的条数")

    # focus
    p2 = sub.add_parser("focus", help="针对单个数据集做定点查询并生成短摘要")
    p2.add_argument("--db", type=str, default=str(DATASET_DB_PATH), help="SQLite DB 路径")
    p2.add_argument("--id", type=int, required=True, help="数据集主键 id")
    p2.add_argument("--user", type=int, required=True, help="调用者 user_uid")
    p2.add_argument("--need", required=True, help="用户本次需求（自然语言）")

    args = p.parse_args(argv)

    if args.cmd == "overview":
        res = await overview(
            user_uid=args.user,
            db_path=Path(args.db),
            limit=args.limit,
        )
    else:
        res = await focus(
            dataset_id=args.id,
            user_need_text=args.need,
            user_uid=args.user,
            db_path=Path(args.db),
        )

    print(json.dumps(res, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    asyncio.run(_amain())
